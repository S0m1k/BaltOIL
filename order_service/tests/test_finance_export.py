"""Юниты генератора финансового XLSX (без БД и без поднятого сервиса).

Запуск из папки order_service:  pytest tests/test_finance_export.py
"""
import io
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.finance_export import finance_payments_xlsx  # noqa: E402

from app.services.finance_export import _COLUMNS  # noqa: E402

_COLUMNS_AMOUNT_INDEX = _COLUMNS.index("Сумма, ₽")



MSK = timezone(timedelta(hours=3))


def _payment(**over):
    base = {
        "payment_id":   "11111111-1111-1111-1111-111111111111",
        "order_number": "ф1",
        "order_kind":   "individual",
        "ttn_number":   "TTN-2026-000042",
        "client_name":  'ООО "ОТК"',
        "kind":         "invoice",
        "payment_type": "postpaid",
        "status":       "paid",
        "method":       "bank_transfer",
        "amount":       67110.5,
        "paid_at":      datetime(2026, 8, 12, 14, 30, tzinfo=MSK),
        "created_at":   datetime(2026, 8, 10, 9, 0, tzinfo=MSK),
        "notes":        "Оплата, частями; перенос",
    }
    base.update(over)
    return base


def _sheet(report):
    wb = load_workbook(io.BytesIO(finance_payments_xlsx(report)))
    return wb.active


def _find_row(ws, first_cell_value):
    for row in ws.iter_rows():
        if row[0].value == first_cell_value:
            return row
    raise AssertionError(f"row {first_cell_value!r} not found")


def _header_row(ws):
    return _find_row(ws, "Дата создания")


def _first_data_row(ws):
    # header, затем строка-заголовок секции, затем данные
    return ws[_header_row(ws)[0].row + 2]


# ── Базовая структура ─────────────────────────────────────────────────────────

def test_produces_real_xlsx_bytes():
    data = finance_payments_xlsx({"payments": []})
    assert data[:2] == b"PK"  # zip-контейнер OOXML


def test_empty_period_still_renders():
    ws = _sheet({"period_from": None, "period_to": None, "payments": []})
    assert ws["A1"].value == "Финансовый отчёт — платежи"
    assert ws["A2"].value == "Период: — — —"
    assert _header_row(ws)[0].value == "Дата создания"


def test_header_columns_order():
    ws = _sheet({"payments": [_payment()]})
    titles = [c.value for c in _header_row(ws) if c.value]
    assert titles[:6] == [
        "Дата создания", "Дата оплаты", "Заявка №", "Вид", "№ ТТН", "Клиент",
    ]


def test_freeze_panes_below_header():
    ws = _sheet({"payments": [_payment()]})
    assert ws.freeze_panes == f"A{_header_row(ws)[0].row + 1}"


# ── Значения ячеек ────────────────────────────────────────────────────────────

def test_amount_is_a_number_not_text():
    ws = _sheet({"payments": [_payment(amount="67110.50")]})
    cell = _first_data_row(ws)[10]
    assert isinstance(cell.value, float)
    assert cell.value == pytest.approx(67110.50)
    assert cell.number_format == "#,##0.00"


def test_dates_are_real_dates_and_tz_stripped():
    ws = _sheet({"payments": [_payment()]})
    created = _first_data_row(ws)[0]
    assert isinstance(created.value, datetime)
    assert created.value.tzinfo is None
    assert created.value == datetime(2026, 8, 10, 9, 0)


def test_iso_strings_are_parsed_into_dates():
    ws = _sheet({"payments": [_payment(created_at="2026-08-10T09:00:00+03:00")]})
    assert _first_data_row(ws)[0].value == datetime(2026, 8, 10, 9, 0)


def test_missing_paid_at_renders_dash():
    ws = _sheet({"payments": [_payment(paid_at=None)]})
    assert _first_data_row(ws)[1].value == "—"


def test_client_name_instead_of_uuid():
    ws = _sheet({"payments": [_payment()]})
    assert _first_data_row(ws)[5].value == 'ООО "ОТК"'


def test_missing_client_name_renders_dash():
    ws = _sheet({"payments": [_payment(client_name=None)]})
    assert _first_data_row(ws)[5].value == "—"


def test_ttn_number_column():
    ws = _sheet({"payments": [_payment()]})
    assert _first_data_row(ws)[4].value == "TTN-2026-000042"
    ws2 = _sheet({"payments": [_payment(ttn_number=None)]})
    assert _first_data_row(ws2)[4].value == "—"


def test_codes_translated_to_russian():
    ws = _sheet({"payments": [_payment()]})
    row = _first_data_row(ws)
    assert row[6].value == "Счёт"                    # kind
    assert row[7].value == "Постоплата (по счёту)"   # payment_type
    assert row[8].value == "Оплачено"                # status
    assert row[9].value == "Банковский перевод"      # method


def test_unknown_code_passes_through():
    ws = _sheet({"payments": [_payment(status="weird_new_status")]})
    assert _first_data_row(ws)[8].value == "weird_new_status"


def test_notes_with_comma_and_newline_stay_in_one_cell():
    # Ровно тот случай, на котором «ехали поля» в CSV.
    ws = _sheet({"payments": [_payment(notes="a,b\nc;d")]})
    assert _first_data_row(ws)[11].value == "a,b\nc;d"


# ── Безопасность ──────────────────────────────────────────────────────────────

@pytest.mark.parametrize("evil", ["=1+1", "+SUM(A1)", "-2", "@cmd", "\tx", "\rx"])
def test_formula_injection_is_neutralised(evil):
    ws = _sheet({"payments": [_payment(notes=evil, client_name=evil)]})
    row = _first_data_row(ws)
    assert row[11].value == "'" + evil
    assert row[5].value == "'" + evil


def test_safe_text_is_not_prefixed():
    ws = _sheet({"payments": [_payment(notes="обычный текст")]})
    assert _first_data_row(ws)[11].value == "обычный текст"


# ── Итоги ─────────────────────────────────────────────────────────────────────

def test_totals_split_paid_and_pending():
    payments = [
        _payment(status="paid", amount=100),
        _payment(status="paid", amount=50.5),
        _payment(status="pending", amount=200),
        _payment(status="cancelled", amount=9999),
    ]
    ws = _sheet({"payments": payments})
    assert _find_row(ws, "Платежей всего")[1].value == 4
    assert _find_row(ws, "Оплачено, ₽")[1].value == pytest.approx(150.5)
    assert _find_row(ws, "Ожидает оплаты, ₽")[1].value == pytest.approx(200.0)


def test_broken_amount_does_not_crash_totals():
    ws = _sheet({"payments": [_payment(amount=None), _payment(amount="—")]})
    assert _find_row(ws, "Оплачено, ₽")[1].value == pytest.approx(0.0)


def test_period_is_formatted_as_russian_dates():
    ws = _sheet({
        "period_from": datetime(2026, 8, 1, tzinfo=MSK),
        "period_to":   "2026-08-31T23:59:59+03:00",
        "payments":    [],
    })
    assert ws["A2"].value == "Период: 01.08.2026 — 31.08.2026"


def test_all_payments_are_rendered():
    payments = [_payment(order_number=f"ф{i + 1}") for i in range(25)]
    ws = _sheet({"payments": payments})
    start = _first_data_row(ws)[0].row
    numbers = [ws.cell(row=start + i, column=3).value for i in range(25)]
    # внутри секции — натуральная сортировка по номеру заявки
    assert numbers == [f"ф{i + 1}" for i in range(25)]


# ── Секции по виду заявки (правки заказчика 2026-09-02) ───────────────────────

def _section_titles(ws):
    """Строки-заголовки секций и подытогов — по первой колонке, ниже шапки."""
    start = _header_row(ws)[0].row
    return [
        ws.cell(row=r, column=1).value
        for r in range(start + 1, ws.max_row + 1)
        if isinstance(ws.cell(row=r, column=1).value, str)
        and (ws.cell(row=r, column=1).value.startswith("Итого")
             or ws.cell(row=r, column=1).value.endswith(")"))
    ]


def test_sections_follow_fixed_order():
    ws = _sheet({"payments": [
        _payment(order_kind="ttn_l",      order_number="л1"),
        _payment(order_kind="company",    order_number="ю1"),
        _payment(order_kind="individual", order_number="ф1"),
    ]})
    titles = _section_titles(ws)
    assert titles == [
        "Физические лица (1)",  "Итого: Физические лица",
        "Юридические лица (1)", "Итого: Юридические лица",
        "ТТН-Л (1)",            "Итого: ТТН-Л",
        "Итого по всем видам",
    ]


def test_empty_sections_are_not_rendered():
    ws = _sheet({"payments": [_payment(order_kind="company")]})
    titles = _section_titles(ws)
    assert "Физические лица (1)" not in titles
    assert titles == ["Юридические лица (1)", "Итого: Юридические лица", "Итого по всем видам"]


def test_kind_filter_leaves_single_section():
    # Как при ?kind=individual: бэкенд отдаёт платежи одного вида.
    ws = _sheet({"payments": [_payment(order_kind="individual") for _ in range(3)]})
    assert _section_titles(ws) == [
        "Физические лица (3)", "Итого: Физические лица", "Итого по всем видам",
    ]


def test_section_subtotals_and_grand_total():
    ws = _sheet({"payments": [
        _payment(order_kind="individual", amount=100),
        _payment(order_kind="individual", amount=50.5),
        _payment(order_kind="company",    amount=200),
    ]})
    amount_col = _COLUMNS_AMOUNT_INDEX
    assert _find_row(ws, "Итого: Физические лица")[amount_col].value == pytest.approx(150.5)
    assert _find_row(ws, "Итого: Юридические лица")[amount_col].value == pytest.approx(200.0)
    assert _find_row(ws, "Итого по всем видам")[amount_col].value == pytest.approx(350.5)


def test_kind_column_shows_short_label():
    ws = _sheet({"payments": [_payment(order_kind="ttn_l")]})
    assert _first_data_row(ws)[3].value == "Л"


def test_unknown_kind_goes_to_other_section_and_keeps_total():
    ws = _sheet({"payments": [
        _payment(order_kind="individual", amount=10),
        _payment(order_kind="", amount=5),
    ]})
    titles = _section_titles(ws)
    assert "Прочие (1)" in titles
    assert _first_data_row(ws)[3].value == "Физ"
    amount_col = _COLUMNS_AMOUNT_INDEX
    assert _find_row(ws, "Итого по всем видам")[amount_col].value == pytest.approx(15.0)


def test_natural_sort_by_order_number_inside_section():
    payments = [
        _payment(order_kind="individual", order_number=n)
        for n in ("ф10", "ф2", "ф1")
    ]
    ws = _sheet({"payments": payments})
    start = _first_data_row(ws)[0].row
    numbers = [ws.cell(row=start + i, column=3).value for i in range(3)]
    assert numbers == ["ф1", "ф2", "ф10"]


def test_same_order_sorted_by_payment_date():
    payments = [
        _payment(order_kind="individual", order_number="ф1",
                 paid_at=datetime(2026, 8, 20, 10, 0, tzinfo=MSK)),
        _payment(order_kind="individual", order_number="ф1",
                 paid_at=datetime(2026, 8, 12, 10, 0, tzinfo=MSK)),
    ]
    ws = _sheet({"payments": payments})
    start = _first_data_row(ws)[0].row
    dates = [ws.cell(row=start + i, column=2).value for i in range(2)]
    assert dates == [datetime(2026, 8, 12, 10, 0), datetime(2026, 8, 20, 10, 0)]
