"""Финансовый отчёт (платежи) в настоящем .xlsx.

Раньше выгрузка была CSV с запятой-разделителем: русский Excel открывает такой
файл одной колонкой («поля поехали»), суммы с точкой считаются текстом, а сырое
примечание с запятой/переводом строки разваливало строку. Теперь — openpyxl,
как в складском отчёте delivery_service.

Модуль намеренно чистый: на вход list[dict], на выход bytes. Никакой БД и
никаких запросов — значит, тестируется юнитом без поднятого сервиса
(tests/test_finance_export.py).
"""
import io
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

__all__ = [
    "PAYMENT_STATUS_RU",
    "PAYMENT_KIND_RU",
    "PAYMENT_METHOD_RU",
    "PAYMENT_TYPE_RU",
    "ORDER_KIND_SECTIONS",
    "ORDER_KIND_SHORT_RU",
    "finance_payments_xlsx",
]

# ── Стили (единые с delivery_service/app/services/excel_service.py) ───────────

_HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
_SUMMARY_FILL = PatternFill("solid", fgColor="D6E4F0")
_PAID_FILL    = PatternFill("solid", fgColor="D6F0D6")
_PENDING_FILL = PatternFill("solid", fgColor="F0E6D6")

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_TITLE_FONT  = Font(name="Calibri", bold=True, size=13, color="1E3A5F")
_LABEL_FONT  = Font(name="Calibri", bold=True, size=10)

_THIN   = Side(style="thin", color="BFBFBF")
_THIN_B = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left", vertical="center")

_MONEY_FMT = "#,##0.00"

# Защита от formula injection: примечание, имя клиента и номер заявки —
# свободный текст. Excel исполнит ячейку, начинающуюся с = + - @.
_XLSX_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _xlsx_safe(value):
    if isinstance(value, str) and value.startswith(_XLSX_FORMULA_TRIGGERS):
        return "'" + value
    return value


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, cols: list[str], row: int) -> None:
    for col, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN_B


def _cell(ws, row: int, col: int, value, fill=None, bold=False, fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=_xlsx_safe(value))
    cell.font      = Font(name="Calibri", bold=bold, size=10)
    cell.alignment = align or _LEFT
    cell.border    = _THIN_B
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# ── Человеческие подписи вместо машинных кодов ───────────────────────────────

PAYMENT_STATUS_RU = {
    "pending":   "Ожидает оплаты",
    "paid":      "Оплачено",
    "cancelled": "Отменено",
}

PAYMENT_KIND_RU = {
    "prepayment": "Предоплата",
    "actual":     "По факту",
    "invoice":    "Счёт",
}

PAYMENT_METHOD_RU = {
    "cash":          "Наличные",
    "card":          "Карта",
    "bank_transfer": "Банковский перевод",
}

# Коды из order_service/app/models/order.py::PaymentType
PAYMENT_TYPE_RU = {
    "prepaid":      "Предоплата",
    "on_delivery":  "По факту (при прибытии)",
    "trade_credit": "Товарный кредит",
    "postpaid":     "Постоплата (по счёту)",
    "debt":         "В долг",
}


# Виды заявок (order_service/app/models/order.py::OrderKind) в порядке секций отчёта.
ORDER_KIND_SECTIONS = (
    ("individual", "Физические лица",  "Физ"),
    ("company",    "Юридические лица", "Юр"),
    ("ttn_l",      "ТТН-Л",            "Л"),
)
ORDER_KIND_SHORT_RU = {code: short for code, _, short in ORDER_KIND_SECTIONS}

# Заявки с неизвестным/пустым видом не выбрасываем: иначе строки платежей молча
# исчезли бы из отчёта, а сумма перестала биться с итогом «Оплачено, ₽».
_OTHER_SECTION = ("Прочие", "—")


def _ru(mapping: dict[str, str], code) -> str:
    """Русская подпись по коду; неизвестный код возвращаем как есть."""
    if code is None or code == "":
        return "—"
    return mapping.get(str(code), str(code))


def _as_dt(value):
    """ISO-строку → datetime; всё остальное отдаём как есть."""
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return value
    return value


def _naive(value):
    """Excel не умеет tz-aware datetime — снимаем таймзону перед записью."""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _fmt_period(value) -> str:
    value = _as_dt(value)
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if value is None:
        return "—"
    return str(value)


def _as_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


# ── Отчёт ────────────────────────────────────────────────────────────────────

_COLUMNS = [
    "Дата создания", "Дата оплаты", "Заявка №", "Вид", "№ ТТН", "Клиент",
    "Вид платежа", "Тип оплаты", "Статус", "Метод", "Сумма, ₽",
    "Примечание", "ID платежа",
]
_COL_WIDTHS = [18, 18, 16, 8, 20, 34, 16, 20, 18, 20, 14, 40, 38]
_LAST_COL = get_column_letter(len(_COLUMNS))
_AMOUNT_COL = _COLUMNS.index("Сумма, ₽") + 1


def _order_number_key(number) -> tuple[int, str]:
    """Натуральная сортировка номеров: «ф9» раньше «ф10» (лексически — наоборот)."""
    text = str(number or "")
    digits = "".join(ch for ch in text if ch.isdigit())
    return (int(digits) if digits else 0, text)


def _payment_sort_key(payment: dict):
    """Внутри секции: по номеру заявки, затем по дате платежа."""
    moment = _naive(_as_dt(payment.get("paid_at") or payment.get("created_at")))
    if not isinstance(moment, datetime):
        moment = datetime.min
    return (*_order_number_key(payment.get("order_number")), moment)


def _sections_by_kind(payments: list[dict]) -> list[tuple[str, list[dict]]]:
    """Разбить платежи на секции по виду заявки; пустые секции не возвращаются."""
    buckets: dict[str, list[dict]] = {code: [] for code, _, _ in ORDER_KIND_SECTIONS}
    other: list[dict] = []
    for p in payments:
        code = str(p.get("order_kind") or "")
        (buckets[code] if code in buckets else other).append(p)

    sections = [
        (title, sorted(buckets[code], key=_payment_sort_key))
        for code, title, _ in ORDER_KIND_SECTIONS
        if buckets[code]
    ]
    if other:
        sections.append((_OTHER_SECTION[0], sorted(other, key=_payment_sort_key)))
    return sections


def finance_payments_xlsx(report: dict) -> bytes:
    """Собрать книгу Excel по платежам.

    report = {
      "period_from": datetime | str | None,
      "period_to":   datetime | str | None,
      "payments": [ {payment_id, order_number, order_kind, ttn_number,
                     client_name, kind, payment_type, status, method, amount,
                     paid_at, created_at, notes}, ... ],
    }

    Платежи выводятся секциями по виду заявки (`order_kind`) с подытогом по
    каждой и общим итогом внизу — заказчик сводит физлиц и юрлиц раздельно.
    """
    payments = report.get("payments") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Финансы"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{_LAST_COL}1")
    title = ws["A1"]
    title.value     = "Финансовый отчёт — платежи"
    title.font      = _TITLE_FONT
    title.alignment = _CENTER

    ws.merge_cells(f"A2:{_LAST_COL}2")
    period = ws["A2"]
    period.value = (
        f"Период: {_fmt_period(report.get('period_from'))}"
        f" — {_fmt_period(report.get('period_to'))}"
    )
    period.font      = Font(name="Calibri", size=10, italic=True)
    period.alignment = _CENTER

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # ── Итоги ────────────────────────────────────────────────────────────────
    total_paid = sum(
        _as_float(p.get("amount")) for p in payments if p.get("status") == "paid"
    )
    total_pending = sum(
        _as_float(p.get("amount")) for p in payments if p.get("status") == "pending"
    )

    kpi_row = 4
    ws.merge_cells(f"A{kpi_row}:{_LAST_COL}{kpi_row}")
    hdr = ws[f"A{kpi_row}"]
    hdr.value     = "Итоги за период"
    hdr.font      = _LABEL_FONT
    hdr.fill      = _SUMMARY_FILL
    hdr.alignment = _LEFT
    hdr.border    = _THIN_B

    kpis = [
        ("Платежей всего", len(payments), None),
        ("Оплачено, ₽",    round(total_paid, 2), _MONEY_FMT),
        ("Ожидает оплаты, ₽", round(total_pending, 2), _MONEY_FMT),
    ]
    for i, (label, value, fmt) in enumerate(kpis):
        r = kpi_row + 1 + i
        _cell(ws, r, 1, label, fill=_SUMMARY_FILL, bold=True)
        ws.merge_cells(f"B{r}:{_LAST_COL}{r}")
        _cell(ws, r, 2, value, fmt=fmt)

    # ── Таблица платежей ─────────────────────────────────────────────────────
    tbl_start = kpi_row + 1 + len(kpis) + 1
    ws.merge_cells(f"A{tbl_start}:{_LAST_COL}{tbl_start}")
    hdr2 = ws[f"A{tbl_start}"]
    hdr2.value     = f"Платежи ({len(payments)})"
    hdr2.font      = _LABEL_FONT
    hdr2.fill      = _SUMMARY_FILL
    hdr2.alignment = _LEFT
    hdr2.border    = _THIN_B

    col_hdr = tbl_start + 1
    _header_row(ws, _COLUMNS, row=col_hdr)

    # Секции по виду заявки (правки заказчика 2026-09-02): физлица, юрлица,
    # ТТН-Л — каждая со своим подытогом, общий итог в конце.
    r = col_hdr
    for section_title, section_payments in _sections_by_kind(payments):
        r += 1
        ws.merge_cells(f"A{r}:{_LAST_COL}{r}")
        sec = ws[f"A{r}"]
        sec.value     = f"{section_title} ({len(section_payments)})"
        sec.font      = _LABEL_FONT
        sec.fill      = _SUMMARY_FILL
        sec.alignment = _LEFT
        sec.border    = _THIN_B

        for p in section_payments:
            r += 1
            status = str(p.get("status") or "")
            fill = _PAID_FILL if status == "paid" else (_PENDING_FILL if status == "pending" else None)

            _cell(ws, r, 1, _naive(_as_dt(p.get("created_at"))), fill=fill, fmt="DD.MM.YYYY HH:MM")
            paid_at = _naive(_as_dt(p.get("paid_at")))
            _cell(ws, r, 2, paid_at if paid_at is not None else "—", fill=fill,
                  fmt="DD.MM.YYYY HH:MM" if paid_at is not None else None)
            _cell(ws, r, 3, p.get("order_number") or "—", fill=fill)
            _cell(ws, r, 4, ORDER_KIND_SHORT_RU.get(str(p.get("order_kind") or ""),
                                                    _OTHER_SECTION[1]), fill=fill)
            _cell(ws, r, 5, p.get("ttn_number") or "—", fill=fill)
            _cell(ws, r, 6, p.get("client_name") or "—", fill=fill)
            _cell(ws, r, 7, _ru(PAYMENT_KIND_RU, p.get("kind")), fill=fill)
            _cell(ws, r, 8, _ru(PAYMENT_TYPE_RU, p.get("payment_type")), fill=fill)
            _cell(ws, r, 9, _ru(PAYMENT_STATUS_RU, status), fill=fill)
            _cell(ws, r, 10, _ru(PAYMENT_METHOD_RU, p.get("method")), fill=fill)
            _cell(ws, r, 11, round(_as_float(p.get("amount")), 2), fill=fill, fmt=_MONEY_FMT)
            _cell(ws, r, 12, p.get("notes") or "", fill=fill)
            _cell(ws, r, 13, str(p.get("payment_id") or ""), fill=fill)

        r += 1
        _cell(ws, r, 1, f"Итого: {section_title}", fill=_SUMMARY_FILL, bold=True)
        _cell(ws, r, _AMOUNT_COL,
              round(sum(_as_float(p.get("amount")) for p in section_payments), 2),
              fill=_SUMMARY_FILL, bold=True, fmt=_MONEY_FMT)

    r += 1
    _cell(ws, r, 1, "Итого по всем видам", fill=_SUMMARY_FILL, bold=True)
    _cell(ws, r, _AMOUNT_COL,
          round(sum(_as_float(p.get("amount")) for p in payments), 2),
          fill=_SUMMARY_FILL, bold=True, fmt=_MONEY_FMT)

    # Шапка таблицы остаётся видимой при прокрутке.
    ws.freeze_panes = ws.cell(row=col_hdr + 1, column=1)
    _set_col_widths(ws, _COL_WIDTHS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
