"""Generate XLSX reports using openpyxl."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Стили ────────────────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
_SUMMARY_FILL  = PatternFill("solid", fgColor="D6E4F0")
_ARRIVAL_FILL  = PatternFill("solid", fgColor="D6F0D6")
_DEPART_FILL   = PatternFill("solid", fgColor="F0E6D6")
_HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_TITLE_FONT    = Font(name="Calibri", bold=True, size=13, color="1E3A5F")
_LABEL_FONT    = Font(name="Calibri", bold=True, size=10)
_NORMAL_FONT   = Font(name="Calibri", size=10)

_THIN   = Side(style="thin",   color="BFBFBF")
_THIN_B = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, cols: list[str], row: int = 1) -> None:
    for col, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN_B


def _cell(ws, row: int, col: int, value, fill=None, bold=False, fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Calibri", bold=bold, size=10)
    cell.alignment = align or _LEFT
    cell.border    = _THIN_B
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# ── Отчёт водителя ───────────────────────────────────────────────────────────

TRIP_STATUS_RU = {
    "planned":    "Запланирован",
    "in_transit": "В пути",
    "completed":  "Завершён",
    "cancelled":  "Отменён",
}


def driver_report_xlsx(report: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт водителя"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Отчёт по рейсам водителя"
    title_cell.font  = _TITLE_FONT
    title_cell.alignment = _CENTER

    ws.merge_cells("A2:G2")
    period = ws["A2"]
    period.value = (
        f"Период: {_fmt(report['period_from'])} — {_fmt(report['period_to'])}"
    )
    period.font      = Font(name="Calibri", size=10, italic=True)
    period.alignment = _CENTER

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # Summary KPIs
    kpi_row = 4
    kpis = [
        ("Всего рейсов",  report["total_trips"]),
        ("Завершено",     report["completed_trips"]),
        ("Отменено",      report["cancelled_trips"]),
        ("Объём план, л", report["total_volume_planned"]),
        ("Объём факт, л", report["total_volume_actual"]),
        ("Пробег, км",    report.get("total_distance_km", "—")),
    ]
    ws.merge_cells(f"A{kpi_row}:G{kpi_row}")
    hdr = ws[f"A{kpi_row}"]
    hdr.value     = "Итоги"
    hdr.font      = _LABEL_FONT
    hdr.fill      = _SUMMARY_FILL
    hdr.alignment = _LEFT
    hdr.border    = _THIN_B

    for i, (label, value) in enumerate(kpis):
        r = kpi_row + 1 + i
        _cell(ws, r, 1, label, fill=_SUMMARY_FILL, bold=True)
        ws.merge_cells(f"B{r}:G{r}")
        val_cell = _cell(ws, r, 2, value)
        val_cell.alignment = _LEFT

    # Trips table
    tbl_start = kpi_row + 1 + len(kpis) + 2
    ws.merge_cells(f"A{tbl_start}:G{tbl_start}")
    hdr2 = ws[f"A{tbl_start}"]
    hdr2.value     = f"Рейсы за период ({len(report['trips'])})"
    hdr2.font      = _LABEL_FONT
    hdr2.fill      = _SUMMARY_FILL
    hdr2.alignment = _LEFT
    hdr2.border    = _THIN_B

    col_hdr = tbl_start + 1
    _header_row(ws, ["Статус", "Адрес", "Объём план", "Объём факт",
                      "Пробег", "Отправление", "Прибытие"], row=col_hdr)

    for i, t in enumerate(report.get("trips", []), 1):
        r = col_hdr + i
        dist = "—"
        if t.get("odometer_start") is not None and t.get("odometer_end") is not None:
            dist = round(float(t["odometer_end"]) - float(t["odometer_start"]), 1)

        fill = _DEPART_FILL if t["status"] == "cancelled" else (
            _ARRIVAL_FILL if t["status"] == "completed" else None
        )
        _cell(ws, r, 1, TRIP_STATUS_RU.get(t["status"], t["status"]), fill=fill)
        _cell(ws, r, 2, t.get("delivery_address", ""), fill=fill)
        _cell(ws, r, 3, float(t.get("volume_planned", 0)), fill=fill, fmt="#,##0.00")
        _cell(ws, r, 4, float(t["volume_actual"]) if t.get("volume_actual") else "—",
              fill=fill, fmt="#,##0.00")
        _cell(ws, r, 5, dist, fill=fill)
        _cell(ws, r, 6, _fmt(t.get("departed_at")), fill=fill)
        _cell(ws, r, 7, _fmt(t.get("arrived_at")), fill=fill)

    _set_col_widths(ws, [14, 40, 14, 14, 10, 18, 18])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Складской отчёт ──────────────────────────────────────────────────────────

TX_TYPE_RU = {"arrival": "Приход", "departure": "Расход"}


def inventory_report_xlsx(report: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Складской отчёт"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Сводный складской отчёт"
    title_cell.font  = _TITLE_FONT
    title_cell.alignment = _CENTER

    ws.merge_cells("A2:H2")
    period_cell = ws["A2"]
    fuel_filter = report.get("fuel_type_filter")
    period_cell.value = (
        f"Период: {_fmt(report['period_from'])} — {_fmt(report['period_to'])}"
        + (f"  |  Вид топлива: {fuel_filter}" if fuel_filter else "")
    )
    period_cell.font      = Font(name="Calibri", size=10, italic=True)
    period_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # Summary table
    row = 4
    ws.merge_cells(f"A{row}:H{row}")
    sh = ws[f"A{row}"]
    sh.value     = "Сводка по видам топлива"
    sh.font      = _LABEL_FONT
    sh.fill      = _SUMMARY_FILL
    sh.alignment = _LEFT
    sh.border    = _THIN_B

    row += 1
    _header_row(ws, ["Вид топлива", "Остаток нач.", "Приход", "Расход", "Остаток кон."],
                row=row)
    for s in report.get("summary", []):
        row += 1
        _cell(ws, row, 1, s["fuel_label"])
        _cell(ws, row, 2, s["opening_balance"], fmt="#,##0.00")
        _cell(ws, row, 3, s["total_arrivals"],  fmt="#,##0.00", fill=_ARRIVAL_FILL)
        _cell(ws, row, 4, s["total_departures"], fmt="#,##0.00", fill=_DEPART_FILL)
        _cell(ws, row, 5, s["closing_balance"],  fmt="#,##0.00")

    # Transactions table
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    dh = ws[f"A{row}"]
    dh.value     = f"Детализация операций ({len(report.get('transactions', []))})"
    dh.font      = _LABEL_FONT
    dh.fill      = _SUMMARY_FILL
    dh.alignment = _LEFT
    dh.border    = _THIN_B

    row += 1
    _header_row(ws, ["Дата", "Тип", "Топливо", "Объём (л)",
                      "Заявка №", "Клиент", "Водитель", "Поставщик / Накладная"],
                row=row)
    for tx in report.get("transactions", []):
        row += 1
        fill = _ARRIVAL_FILL if tx["type"] == "arrival" else _DEPART_FILL
        _cell(ws, row, 1, _fmt(tx.get("transaction_date")),   fill=fill)
        _cell(ws, row, 2, TX_TYPE_RU.get(tx["type"], tx["type"]), fill=fill)
        _cell(ws, row, 3, tx.get("fuel_label", tx.get("fuel_type", "")), fill=fill)
        _cell(ws, row, 4, float(tx.get("volume", 0)), fill=fill, fmt="#,##0.00")
        _cell(ws, row, 5, tx.get("order_number") or "—", fill=fill)
        _cell(ws, row, 6, tx.get("client_name") or "—",  fill=fill)
        _cell(ws, row, 7, tx.get("driver_name") or "—",  fill=fill)
        supplier = tx.get("supplier_name") or ""
        invoice  = tx.get("invoice_number") or ""
        _cell(ws, row, 8, f"{supplier} {invoice}".strip() or "—", fill=fill)

    _set_col_widths(ws, [18, 10, 16, 14, 16, 22, 22, 24])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt(val) -> str:
    if val is None:
        return "—"
    if isinstance(val, str):
        try:
            val = datetime.fromisoformat(val.replace("Z", "+00:00"))
        except ValueError:
            return val
    if isinstance(val, datetime):
        return val.strftime("%d.%m.%Y %H:%M")
    return str(val)
