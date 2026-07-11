import uuid
from datetime import datetime, date
from io import BytesIO
from typing import Annotated
from fastapi import APIRouter, Depends, Request, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.models.user import User, UserRole
from app.models.client_profile import ClientProfile, ClientType
from app.models.audit_log import AuditLog
from app.core.dependencies import CurrentUser, require_roles, get_request_meta
from app.core.rate_limit import limiter
from app.schemas.auth import ChangePasswordRequest
from app.schemas.user import UserResponse, UserShortResponse, UserDirectoryEntry, CreateUserRequest, UpdateUserRequest, ClientExportRequest
from app.schemas.client_profile import ClientProfileResponse, UpdateClientProfileRequest, UpdateClientTariffRequest
from app.services import user_service
from app.services.audit_service import log_action

router = APIRouter(prefix="/users", tags=["users"])

AdminOnly = Annotated[object, Depends(require_roles(UserRole.ADMIN))]
AdminOrManager = Annotated[object, Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER))]
# Разовых клиентов может заводить и водитель (создание заявки в мобильном приложении)
StaffOrDriver = Annotated[object, Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER, UserRole.DRIVER))]


@router.get("", response_model=list[UserShortResponse])
async def list_users(
    _: AdminOrManager,
    db: Annotated[AsyncSession, Depends(get_db)],
    role: UserRole | None = Query(None),
    include_inactive: bool = Query(False),
    offset: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    client_number: int | None = Query(None, description="Найти клиента по короткому номеру"),
    one_off: bool | None = Query(None, description="Только разовые клиенты (true) / только обычные (false)"),
):
    users = await user_service.list_users(
        db,
        role=role,
        include_inactive=include_inactive,
        offset=offset,
        limit=limit,
        client_number=client_number,
        one_off=one_off,
    )
    result = []
    for u in users:
        entry = UserShortResponse.model_validate(u)
        # Read from instance dict to avoid triggering SQLAlchemy lazy load
        profile = u.__dict__.get("client_profile")
        if profile is not None:
            entry.client_number = profile.client_number
            entry.is_one_off = profile.is_one_off
        result.append(entry)
    return result


@router.get("/directory", response_model=list[UserDirectoryEntry])
async def users_directory(
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    role: UserRole | None = Query(None, description="Фильтр по роли"),
    limit: int = Query(500, ge=1, le=1000),
):
    """Адресная книга: id + ФИО + роль активных пользователей.
    Доступна любому залогиненному (нужна чату для резолва UUID → имя), но
    клиенту никогда не видны другие клиенты — это утечка клиентской базы.
    Клиент видит только staff (admin/manager/driver). Не возвращает email/телефон.
    """
    if current_user.role == UserRole.CLIENT:
        # Клиент может фильтровать только по staff-ролям. Если запросил CLIENT —
        # вернём пустой список вместо 403, чтобы клиент UI не падал.
        if role == UserRole.CLIENT:
            return []
        # Исключаем клиентов на уровне SQL (см. B9): иначе LIMIT отрезал бы staff.
        return await user_service.list_users(
            db, role=role, exclude_role=UserRole.CLIENT,
            include_inactive=False, offset=0, limit=limit,
        )
    return await user_service.list_users(
        db, role=role, include_inactive=False, offset=0, limit=limit
    )


class OneOffClientRequest(BaseModel):
    full_name: str
    phone: str


@router.post("/one-off", response_model=UserShortResponse)
async def create_one_off_client(
    data: OneOffClientRequest,
    current_user: CurrentUser,
    _: StaffOrDriver,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Разовый клиент (правки 2026-07-11): физлицо, обязательны имя и телефон.

    Идемпотентен по телефону: если клиент с таким номером уже есть (разовый или
    зарегистрированный) — возвращается существующий, дубли не создаются.
    """
    meta = get_request_meta(request)
    user = await user_service.get_or_create_one_off_client(
        db,
        full_name=data.full_name,
        phone=data.phone,
        actor_id=current_user.id,
        ip_address=meta["ip_address"],
    )
    entry = UserShortResponse.model_validate(user)
    profile = user.__dict__.get("client_profile")
    if profile is not None:
        entry.client_number = profile.client_number
        entry.is_one_off = profile.is_one_off
    return entry


@router.post("", response_model=UserResponse, status_code=201)
async def create_user(
    data: CreateUserRequest,
    current_user: CurrentUser,
    _: AdminOnly,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    meta = get_request_meta(request)
    return await user_service.create_user_by_admin(
        db, data, actor_id=current_user.id, ip_address=meta["ip_address"]
    )


@router.get("/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    # Clients can only see themselves
    if current_user.role == UserRole.CLIENT and current_user.id != user_id:
        from app.core.exceptions import ForbiddenError
        raise ForbiddenError()
    return await user_service.get_user_by_id(db, user_id)


@router.patch("/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: uuid.UUID,
    data: UpdateUserRequest,
    current_user: CurrentUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    meta = get_request_meta(request)
    return await user_service.update_user(
        db, user_id, data, actor=current_user, ip_address=meta["ip_address"]
    )


@router.delete("/{user_id}", status_code=204)
async def archive_user(
    user_id: uuid.UUID,
    _: AdminOnly,
    current_user: CurrentUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    meta = get_request_meta(request)
    await user_service.archive_user(
        db, user_id, actor=current_user, ip_address=meta["ip_address"]
    )


@router.patch("/{user_id}/profile", response_model=ClientProfileResponse)
@limiter.limit("30/minute")
async def update_client_profile(
    user_id: uuid.UUID,
    data: UpdateClientProfileRequest,
    current_user: CurrentUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    meta = get_request_meta(request)
    return await user_service.update_client_profile(
        db, user_id, data, actor=current_user, ip_address=meta["ip_address"]
    )


@router.patch("/{user_id}/tariff", response_model=ClientProfileResponse)
async def update_client_tariff(
    user_id: uuid.UUID,
    data: UpdateClientTariffRequest,
    current_user: CurrentUser,
    _: AdminOnly,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    return await user_service.update_client_tariff(
        db, user_id, data, actor=current_user
    )


@router.post("/{user_id}/fns-resync", response_model=UserResponse)
@limiter.limit("10/minute")
async def fns_resync(
    user_id: uuid.UUID,
    current_user: CurrentUser,
    _: AdminOrManager,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Перетянуть реквизиты юрлица из DaData (ЕГРЮЛ + банк-справочник)
    и обновить ClientProfile. Доступ: admin / manager.
    """
    meta = get_request_meta(request)
    return await user_service.fns_resync_user(
        db, user_id, actor_id=current_user.id, ip_address=meta["ip_address"]
    )


_XLSX_COLUMNS = [
    "Тип", "Наименование", "ИНН", "КПП", "ОГРН", "БИК", "Банк", "Р/с",
    "Корр.счёт", "Юр.адрес", "Адрес доставки", "Тариф", "Credit allowed",
    "Credit limit", "Email логин", "Billing email", "Телефон", "Дата регистрации",
]
_XLSX_COL_WIDTHS = [12, 40, 14, 10, 16, 11, 35, 22, 22, 45, 45, 38, 14, 14, 32, 32, 18, 20]

# Excel/LibreOffice трактуют строки, начинающиеся с этих символов, как формулу.
# Клиент с company_name '=cmd|...' заразит файл, который потом скачает менеджер.
# Префиксируем одинарной кавычкой — она съедается при отображении, но рвёт парсер формул.
_XLSX_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _xlsx_safe(value):
    """Защита от CSV/formula injection. Числа и None — без изменений."""
    if isinstance(value, str) and value.startswith(_XLSX_FORMULA_TRIGGERS):
        return "'" + value
    return value


@router.post("/clients/export")
async def export_clients(
    data: ClientExportRequest,
    current_user: CurrentUser,
    _: AdminOrManager,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> StreamingResponse:
    """Export selected clients to xlsx. Manager/admin only."""
    # Lazy import: openpyxl нужен только здесь. Если его нет в образе
    # (старый build до sprint 2026-06 deploy 4) — top-level import уронит весь
    # сервис. Сейчас импортим только при вызове, отдаём 503 если пакета нет.
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(status_code=503, detail="xlsx export module not installed; rebuild auth_service image")

    if len(data.client_ids) > 1000:
        raise HTTPException(status_code=400, detail="too many: max 1000 clients per export")

    result = await db.execute(
        select(User)
        .options(selectinload(User.client_profile))
        .where(User.id.in_(data.client_ids), User.role == UserRole.CLIENT)
    )
    users = list(result.scalars().all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    bold = Font(bold=True)
    for col_idx, (header, width) in enumerate(zip(_XLSX_COLUMNS, _XLSX_COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, u in enumerate(users, start=2):
        p: ClientProfile | None = u.client_profile
        is_company = p is not None and p.client_type == ClientType.COMPANY

        client_type_label = "Юр.лицо" if is_company else "Физлицо"
        name = (p.company_name if is_company and p else None) or u.full_name
        inn = p.inn if p else None
        kpp = p.kpp if p else None
        ogrn = p.ogrn if p else None
        bik = p.bik if p else None
        bank_name = p.bank_name if p else None
        bank_account = p.bank_account if p else None
        correspondent_account = p.correspondent_account if p else None
        legal_address = p.legal_address if p else None
        delivery_address = p.delivery_address if p else None
        tariff_id = str(p.tariff_id) if p and p.tariff_id else None
        credit_allowed = "Да" if p and p.credit_allowed else "Нет"
        credit_limit = float(p.credit_limit) if p and p.credit_limit is not None else None
        billing_email = p.billing_email if p else None
        # Format datetime without timezone suffix
        created_at = u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else None

        row_data = [
            client_type_label, name, inn, kpp, ogrn, bik, bank_name, bank_account,
            correspondent_account, legal_address, delivery_address, tariff_id,
            credit_allowed, credit_limit, u.email, billing_email, u.phone, created_at,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_xlsx_safe(value))

    meta = get_request_meta(request)
    await log_action(
        db,
        action="clients.exported",
        actor_id=current_user.id,
        details={"count": len(users), "client_ids": [str(cid) for cid in data.client_ids]},
        ip_address=meta["ip_address"],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"clients_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class AuditLogResponse(BaseModel):
    id: uuid.UUID
    actor_id: uuid.UUID | None
    action: str
    entity_type: str | None
    entity_id: str | None
    details: dict | None
    ip_address: str | None
    created_at: datetime

    model_config = {"from_attributes": True}


@router.get("/{user_id}/audit", response_model=list[AuditLogResponse])
async def get_user_audit(
    user_id: uuid.UUID,
    _: AdminOrManager,
    db: Annotated[AsyncSession, Depends(get_db)],
    limit: int = Query(50, ge=1, le=200),
):
    """Last N audit_log entries where actor_id or entity_id matches this user. Manager/admin only."""
    result = await db.execute(
        select(AuditLog)
        .where(
            (AuditLog.actor_id == user_id) |
            (AuditLog.entity_id == str(user_id))
        )
        .order_by(AuditLog.created_at.desc())
        .limit(limit)
    )
    return list(result.scalars().all())


@router.post("/me/change-password", status_code=204)
async def change_password(
    data: ChangePasswordRequest,
    current_user: CurrentUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    meta = get_request_meta(request)
    await user_service.change_password(
        db,
        actor=current_user,
        current_password=data.current_password,
        new_password=data.new_password,
        ip_address=meta["ip_address"],
    )
